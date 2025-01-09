import openpyxl
import yaml
import os
import shutil
from datetime import datetime

# Load the Excel file
excel_file = 'configurations.xlsx'
wb = openpyxl.load_workbook(excel_file)

# Load the master data from Sheet1
sheet_master = wb['Sheet1']
master_headers = [cell.value for cell in sheet_master[1]]

# Iterate over all rows in the master sheet
for master_row in sheet_master.iter_rows(min_row=2, values_only=True):
    master_config = {master_headers[i]: cell for i, cell in enumerate(master_row)}

    # Load the Argo data from Sheet2
    sheet_argo = wb['Sheet2']
    argo_headers = [cell.value for cell in sheet_argo[1]]

    # Iterate over all rows in the Argo sheet
    for argo_row in sheet_argo.iter_rows(min_row=2, values_only=True):
        argo_config = {argo_headers[i]: cell for i, cell in enumerate(argo_row)}

        # Load the first YAML template
        template_file_1 = 'template1.yml'
        with open(template_file_1, 'r') as file:
            yaml_content_1 = yaml.safe_load_all(file)
            yaml_docs_1 = list(yaml_content_1)

        # Update the first YAML content with values from the configurations
        for doc in yaml_docs_1:
            if doc['kind'] == 'AppProject':
                doc['metadata']['name'] = master_config['namespace']
                doc['spec']['destinations'][0]['namespace'] = master_config['namespace']
            elif doc['kind'] == 'Application':
                doc['metadata']['name'] = master_config['repo_name']
                doc['spec']['destination']['namespace'] = master_config['namespace']
                doc['sources'][0]['repoURL'] = 'https://git.delta.com/crewdevops/crew-helm-templates.git'
                doc['sources'][0]['targetRevision'] = argo_config['targetRevision']

                # Update the helm valueFiles based on targetRevision
                value_files = argo_config['valueFiles'].split(',')
                doc['sources'][0]['helm']['valueFiles'] = [vf.strip() for vf in value_files]

                doc['sources'][1]['repoURL'] = master_config['app_repo_url']
                doc['sources'][1]['targetRevision'] = argo_config['targetRevision']
                doc['project'] = master_config['namespace']

        # Determine the output file path for the first YAML
        output_file_1 = os.path.join(argo_config['file_path'], argo_config['output_file'])

        # Ensure the directory exists
        os.makedirs(argo_config['file_path'], exist_ok=True)

        # Check if the output file already exists
        if os.path.exists(output_file_1):
            # Create a backup directory if it doesn't exist
            backup_dir = 'c:/tmp'
            os.makedirs(backup_dir, exist_ok=True)

            # Generate a timestamp
            timestamp = datetime.now().strftime('%Y%m%d%H%M%S%f')

            # Move the existing file to the backup directory
            backup_file = os.path.join(backup_dir, f"{os.path.basename(output_file_1).split('.')[0]}_{timestamp}.yml")
            shutil.move(output_file_1, backup_file)
            print(f"Existing file moved to {backup_file}")

        # Save the updated YAML content to the specified output file
        with open(output_file_1, 'w') as file:
            yaml.dump_all(yaml_docs_1, file)

        print(f"Updated YAML file saved as {output_file_1}")

    # Load the valueFile data from Sheet3
    sheet_valuefile = wb['Sheet3']
    valuefile_headers = [cell.value for cell in sheet_valuefile[1]]

    # Iterate over all rows in the valueFile sheet
    for valuefile_row in sheet_valuefile.iter_rows(min_row=2, values_only=True):
        valuefile_config = {valuefile_headers[i]: cell for i, cell in enumerate(valuefile_row)}

        # Load the second YAML template
        template_file_2 = 'template2.yml'
        with open(template_file_2, 'r') as file:
            yaml_content_2 = yaml.safe_load(file)

        # Update the second YAML content with values from the configurations
        yaml_content_2['serviceAccount'][0]['name'] = f"{master_config['namespace']}-{master_config['repo_name']}-app"
        yaml_content_2['serviceAccount'][0]['annotations']['eks.amazonaws.com/role-arn'] = (
            f"arn:aws:iam ::{valuefile_config['arn_number']}:role/delegate-admin-{master_config['namespace']}-{valuefile_config['region']}-deployer-role"
        )
        yaml_content_2['routeValues']['host'] = f"{master_config['repo_name']}-{valuefile_config['env']}.{valuefile_config['cluster-name']}-{valuefile_config['env']}.aws.delta.com"
        yaml_content_2['routeValues']['tls']['privateKeySecret'] = f"{master_config['repo_name']}-tls-private"
        yaml_content_2['routeValues']['tls']['publicKeySecret'] = f"{master_config['repo_name']}-tls-public"
        yaml_content_2['env'][0]['value'] = valuefile_config['env']

        # Determine the output file path for the second YAML
        output_file_2 = os.path.join(valuefile_config['file_path'], valuefile_config['output_file'])

        # Ensure the directory exists
        os.makedirs(valuefile_config['file_path'], exist_ok=True)

        # Check if the output file already exists
        if os.path.exists(output_file_2):
            # Create a backup directory if it doesn't exist
            backup_dir = 'c:/tmp'
            os.makedirs(backup_dir, exist_ok=True)

            # Generate a timestamp
            timestamp = datetime.now().strftime('%Y%m%d%H%M%S%f')

            # Move the existing file to the backup directory
            backup_file = os.path.join(backup_dir, f"{os.path.basename(output_file_2).split('.')[0]}_{timestamp}.yml")
            shutil.move(output_file_2, backup_file)
            print(f"Existing file moved to {backup_file}")

        # Save the updated YAML content to the specified output file
        with open(output_file_2, 'w') as file:
            yaml.dump(yaml_content_2, file)

        print(f"Updated YAML file saved as {output_file_2}")
