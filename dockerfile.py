import openpyxl
import os

# Load the Excel file
excel_file = 'configurations.xlsx'
wb = openpyxl.load_workbook(excel_file)
sheet1 = wb['Sheet1']
sheet4 = wb['Sheet4']

# Load the Dockerfile template from a file
with open('dockerfile_template', 'r') as template_file:
    dockerfile_template = template_file.read()

# Get headers from the first row of Sheet1
headers_sheet1 = [cell.value for cell in sheet1[1]]

# Get headers from the first row of Sheet4
headers_sheet4 = [cell.value for cell in sheet4[1]]

# Iterate over each row in Sheet1 starting from the second row
for row in sheet1.iter_rows(min_row=2, values_only=True):
    # Create a dictionary with header names as keys and cell values as values for Sheet1
    row_data_sheet1 = {headers_sheet1[i]: cell for i, cell in enumerate(row)}

    # Get the block_code from the current row in Sheet1
    block_code = row_data_sheet1.get('block_code')

    # Find the matching row in Sheet4 based on block_code
    matching_row_data_sheet4 = None
    for row in sheet4.iter_rows(min_row=2, values_only=True):
        row_data_sheet4 = {headers_sheet4[i]: cell for i, cell in enumerate(row)}
        if row_data_sheet4.get('block_code') == block_code:
            matching_row_data_sheet4 = row_data_sheet4
            break

    if matching_row_data_sheet4:
        # Replace placeholders in the Dockerfile template with actual values from Sheet1 and Sheet4
        dockerfile_content = dockerfile_template
        for key, value in {**row_data_sheet1, **matching_row_data_sheet4}.items():
            placeholder = f"<{key}>"
            dockerfile_content = dockerfile_content.replace(placeholder, str(value) if value is not None else "")

        # Get the file path from the current row in Sheet1
        dockerfile_path = row_data_sheet1.get('dockerfilepath')

        # Ensure the directory exists
        if dockerfile_path:
            os.makedirs(os.path.dirname(dockerfile_path), exist_ok=True)

            # Write the updated Dockerfile content to the specified path
            with open(dockerfile_path, 'w') as dockerfile:
                dockerfile.write(dockerfile_content)

            print(f"Updated Dockerfile saved at {dockerfile_path}")
    else:
        print(f"No matching block_code found in Sheet4 for block_code: {block_code}")
