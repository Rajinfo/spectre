import openpyxl
import os
import json

# Load the Excel file
excel_file = 'configurations.xlsx'
wb = openpyxl.load_workbook(excel_file)
sheet1 = wb['Sheet1']
sheet5 = wb['Sheet5']

# Load the JSON template from a file
with open('json_template.json', 'r') as json_template_file:
    json_template = json.load(json_template_file)

# Get headers from the first row of Sheet1
headers_sheet1 = [cell.value for cell in sheet1[1]]

# Get headers from the first row of Sheet5
headers_sheet5 = [cell.value for cell in sheet5[1]]

# Iterate over each row in Sheet1 starting from the second row
for row in sheet1.iter_rows(min_row=2, values_only=True):
    # Create a dictionary with header names as keys and cell values as values for Sheet1
    row_data_sheet1 = {headers_sheet1[i]: cell for i, cell in enumerate(row)}

    # Check if the status is TRUE
    if row_data_sheet1.get('status') == True:
        # Get the syscode and repo_name from the current row in Sheet1
        syscode = row_data_sheet1.get('syscode')
        repo_name = row_data_sheet1.get('repo_name')

        # Find all matching rows in Sheet5 based on syscode
        matching_rows_sheet5 = []
        for row in sheet5.iter_rows(min_row=2, values_only=True):
            row_data_sheet5 = {headers_sheet5[i]: cell for i, cell in enumerate(row)}
            if row_data_sheet5.get('syscode') == syscode:
                matching_rows_sheet5.append(row_data_sheet5)

        # Iterate over each matching row in Sheet5
        for matching_row in matching_rows_sheet5:
            # Create a copy of the JSON template
            json_data = json_template.copy()

            # Update JSON data with values from Sheet1
            json_data.update({
                "block_code": row_data_sheet1.get('block_code'),
                "namespace": row_data_sheet1.get('namespace'),
                "gitlab_pipeline_project": row_data_sheet1.get('gitlab_pipeline_project'),
                "user_story_ids": row_data_sheet1.get('user_story_ids')
            })

            # Update JSON data with values from Sheet5 only if the key exists in the JSON template
            for key in json_template.keys():
                if key in row_data_sheet5:
                    json_data[key] = row_data_sheet5[key]

            # Construct the file path
            file_location = matching_row.get('file_location')
            filename = matching_row.get('filename')
            if file_location and filename:
                file_path = os.path.join(repo_name, file_location, filename)

                # Ensure the directory exists
                os.makedirs(os.path.dirname(file_path), exist_ok=True)

                # Write the JSON data to the file
                with open(file_path, 'w') as json_file:
                    json.dump(json_data, json_file, indent=4)

                print(f"JSON file created at {file_path}")
    else:
        print(f"Skipping row with syscode {row_data_sheet1.get('syscode')} as status is not TRUE")
